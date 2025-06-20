import openpyxl as xl
import os
from init.app_logger import *
from init.lib import *

#init
template_file = 'template/gfk_template.xlsx'
input_folder = 'input'
start_row = 5

# open the workbook
wb = xl.load_workbook(template_file)
ws = wb.active

# get all file .xlsx in input folder
input_files = os.listdir(input_folder)
input_files = [file for file in input_files if '~$' not in file]
input_files.remove('.gitkeep')

logger.debug(f'Total {len(input_files)} files in input folder')

master_wb_current_row = start_row
for input_file in input_files:
    logger.debug(f'Processing: {input_file}')
    # open the input file
    input_file_path = os.path.join(input_folder, input_file)
    input_wb = xl.load_workbook(input_file_path, data_only=True)
    input_ws = input_wb.active

    # get total rows of input file
    total_rows = input_ws.max_row

    # loop from start_row to total_rows
    for row in range(start_row, total_rows + 1):
        # get the value of the first cell in the row
        dealer_name = input_ws.cell(row, 4).value
        if dealer_name:
            start_col = 6
            
            # branch
            ws.cell(master_wb_current_row, 2).value = input_ws.cell(row, 2).value

            # department
            ws.cell(master_wb_current_row, 3).value = input_ws.cell(row, 3).value

            # dealer_name
            ws.cell(master_wb_current_row, 4).value = dealer_name
            
            # dealer code
            ws.cell(master_wb_current_row, 5).value = input_ws.cell(row, 5).value

            # province
            ws.cell(master_wb_current_row, 6).value = input_ws.cell(row, 6).value

            for quater_start_col in [1, 41, 91 ]:
                for month in [0, 1, 2]:
                    # capacity_previous 
                    capacity_previous_col = start_col + quater_start_col + month*10
                    copy_cell_value(ws, master_wb_current_row, capacity_previous_col,
                                    input_ws, row, capacity_previous_col,
                                    logger, input_file)       

                    # capacity_current
                    capacity_current_col = capacity_previous_col + 1
                    copy_cell_value(ws, master_wb_current_row, capacity_current_col,
                                    input_ws, row, capacity_current_col,
                                    logger, input_file)

                    # dk_previous
                    dk_previous_col = capacity_previous_col + 2
                    copy_cell_value(ws, master_wb_current_row, dk_previous_col,
                                    input_ws, row, dk_previous_col,
                                    logger, input_file)
                    
                    # dk_current
                    dk_current_col = capacity_current_col + 3
                    copy_cell_value(ws, master_wb_current_row, dk_current_col,
                                    input_ws, row, dk_current_col,
                                    logger, input_file)
                    
                    # pn_prevous
                    pn_previous_col = dk_previous_col + 6
                    copy_cell_value(ws, master_wb_current_row, pn_previous_col,
                                    input_ws, row, pn_previous_col,
                                    logger, input_file)
                    
                    # pn_current
                    pn_current_col = dk_current_col + 7
                    copy_cell_value(ws, master_wb_current_row, pn_current_col,
                                    input_ws, row, pn_current_col,
                                    logger, input_file)

            master_wb_current_row += 1

# delete empty rows
ws.delete_rows(master_wb_current_row, ws.max_row - master_wb_current_row + 1)
export_file_name = f'output/{timestamp} Gfk Dealer Monitor (combined).xlsx'
wb.save(export_file_name)
logger.debug(f'Exported: {export_file_name}')
logger.debug('Script Done!')